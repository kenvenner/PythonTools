"""
@author:   Ken Venner
@contact:  ken@venerllc.com
@version:  1.08

Library of tools to work directly with Excel files
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os

# global variables
AppVersion = '1.08'

def open_xlsx_get_ws_wb( xls_filename, ws_sheetname = None, disp_msg=False ):
    """
    Pass in a filename
    Open a xlsx document to the specified sheet name or the active sheetname
    """
    wb = openpyxl.load_workbook( xls_filename )
    if ws_sheetname:
        ws = wb[ws_sheetname]
    else:
        ws = wb.active
    return ws, wb

def get_existing_column_width( xls_filename, ws_sheetname = None, disp_msg=False ):
    """
    Pass in a filename
    Extract the column widths from a define xlsx filename
    """
    col_width = {}
    if not os.path.exists( xls_filename ):
        if disp_msg:
            print('get_existing_column_width:file did not exist for:', xls_filename)
        return col_width
    ws, wb = open_xlsx_get_ws_wb( xls_filename, ws_sheetname, disp_msg=disp_msg )
    for k, cd in ws.column_dimensions.items():
        col_width[k] = cd.width
    return col_width

def get_existing_column_hidden( xls_filename, ws_sheetname = None, disp_msg=False ):
    """
    Pass in a filename
    Extract the column widths from a define xlsx filename
    """
    col_hidden = []
    if not os.path.exists( xls_filename ):
        if disp_msg:
            print('get_existing_column_hidden:file did not exist for:', xls_filename)
        return col_hidden

    # open file 
    ws, wb = open_xlsx_get_ws_wb( xls_filename, ws_sheetname, disp_msg=disp_msg )

    # Determine the maximum column number in the sheet
    max_col = ws.max_column

    # Create a list of all column letters in the sheet (e.g., ['A', 'B', 'C', ...])
    all_cols_letters = [get_column_letter(i) for i in range(1, max_col + 1)]

    col_hidden = []
    last_hidden_group_max = 0

    for i, col_letter in enumerate(all_cols_letters):
        # Get the column dimension object (openpyxl creates it if it doesn't exist)
        col_dim = ws.column_dimensions[col_letter]
    
        # Check if the column is marked as hidden
        if col_dim.hidden:
            col_hidden.append(col_letter)
            # If it's part of a group, record the end of the group
            if col_dim.max is not None:
                last_hidden_group_max = col_dim.max
        # Check if the current column falls within a previously identified hidden group
        elif i + 1 <= last_hidden_group_max:
            col_hidden.append(col_letter)
        # If not hidden or in a group, reset the group tracker
        else:
            last_hidden_group_max = 0

    return col_hidden

# convert this into a class and then apply to that object you opened
def apply_col_width_ws_obj( ws, col_width, disp_msg=True ):
    """
    Pass in a worksheet object and a column width dictionary
    Format the worksheet object to have column widths as defined in the dictionary
    """
    if list(ws.column_dimensions):
        # print('column_dimensions')
        for k, cd in ws.column_dimensions.items():
            if k in col_width and col_width[k]:
                # print('applied width to %s: %d', k, col_width[k])
                ws.column_dimensions[k].width = col_width[k]
            else:
                if disp_msg:
                    print('Skipped column width change - new width not defined: ', k)
    else:
        # print('column range')
        for col in range(ws.min_column, ws.max_column):
            k = get_column_letter(col)
            if k in col_width and col_width[k]:
                # print('applied width to %s: %d', k, col_width[k])
                ws.column_dimensions[k].width = col_width[k]
            else:
                if disp_msg:
                    print('Skipped column: ', k)
            
# convert this into a class and then apply to that object you opened
def apply_col_hidden_ws_obj( ws, col_hidden, disp_msg=True ):
    """
    Pass in a worksheet object and a column width dictionary
    Format the worksheet object to have column widths as defined in the dictionary
    """
    for col in col_hidden:
        ws.column_dimensions[col].hidden = True
        
            
def apply_filter_all_columns( ws ):
    """
    Pass in a worksheet option
    Enble the filter all columns feature with row 1 as the header
    """
    ws.auto_filter.ref = ws.dimensions

def apply_row_freeze( ws, cell='A2'):
    """
    Pass in a worksheet object, and optoinally a position
    Cause this worksheet to freeze to the top row of the row/col passed in
    """
    ws.freeze_panes = cell

def apply_row_bold( ws, row=1):
    """
    Pass in a worksheet object, and optoinally a position
    Cause this worksheet to apply the bold font to the row defined
    """
    # https://www.javatpoint.com/python-openpyxl#IteratebyColumn
    for row in ws.iter_rows(min_row=row, max_row=row):
        for cell in row:
            cell.font = Font(bold=True)



def autofit_column_width( ws ):
    """
    Pass in a work sheet object
    Calculate the column widths to auto-fit the column to the data
    """
    # https://gist.github.com/summerofgeorge/96dac94293b60c70d11d7cd7e852ffd6
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width


def format_xlsx_with_filter_and_freeze( xls_filename, ws_sheetname=None, col_width=None, col_hidden=None, disp_msg=True ):
    """
    Take in a xlsx filename
    Open this file
    If col_width is not None - then format columns to the defined with, if None, autofit the column width
    if col_hidden is not None - then hide the columns defined in this list
    Filter
    Freeze
    Save back to the original filename
    """
    if not os.path.exists( xls_filename ):
        raise Exception('File does not exist: ' + xls_filename)
    ws, wb = open_xlsx_get_ws_wb(xls_filename, disp_msg=disp_msg)
    if col_width:
        if disp_msg:
            print('Applying defined col_width')
        apply_col_width_ws_obj(ws, col_width, disp_msg=disp_msg)
    else:
        if disp_msg:
            print('Autofit col_width')
        autofit_column_width(ws)
    if col_hidden:
        apply_col_hidden_ws_obj(ws, col_hidden, disp_msg=disp_msg)
    apply_filter_all_columns(ws)
    apply_row_freeze(ws)
    # save out the filename
    wb.save(xls_filename)
    
       

# when not used as a library
if __name__ == '__main__':
    src_dir = '/Users/116919/Sierra Space Corporation/Information Technology - PO_Files/'
    src_fname = 'SierraSpace-Open-PO-Lines.xlsx'
    dst_dir = '/Users/116919/Sierra Space Corporation/Information Technology - PO_Files/po-requestor/'
    dst_fname = 'Abraham_Rez_SierraSpace-Open-PO-Lines.xlsx'
    dst_fname2 = 'Abraham_Rez_SierraSpace-Open-PO-Lines_fmt.xlsx'
    dst_fname3 = 'Abraham_Rez_SierraSpace-Open-PO-Lines_fmt2.xlsx'


    # get the master file column widths
    col_width = get_existing_column_width(src_dir+src_fname)

    # get the to be worked on Excel workbook and worksheet
    ws, wb = open_xlsx_get_ws_wb(dst_dir+dst_fname)

    # save for the other test
    wb.save(dst_dir+dst_fname3)

    ## TEST1
    
    # step through and format
    # apply_col_width_ws_obj(ws, col_width)
    autofit_column_width( ws )
    apply_filter_all_columns(ws)
    apply_row_freeze(ws)

    # save out the filename
    wb.save(dst_dir+dst_fname2)
    

    ## TEST2
    format_xlsx_with_filter_and_freeze( dst_dir+dst_fname3, col_width=col_width )
#eof
