import kvxls
import kvutil
import unittest
import datetime
import pprint
from openpyxl.styles import PatternFill
import openpyxl

# logging
import kvlogger


"""
test to add 

"""

config = kvlogger.get_config("t_kvxls.log", loggerlevel="DEBUG")
kvlogger.dictConfig(config)
logger = kvlogger.getLogger(__name__)

# set up filenames
filenamexls = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest",
        "file_ext": ".xls",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)
filenamexlsx = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest",
        "file_ext": ".xlsx",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)

filenamexls2 = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest2",
        "file_ext": ".xls",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)
filenamexlsx2 = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest2",
        "file_ext": ".xlsx",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)

filenamexls3 = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest3",
        "file_ext": ".xls",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)
filenamexlsx3 = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest3",
        "file_ext": ".xlsx",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)

filenamexls4 = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest4",
        "file_ext": ".xls",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)
filenamexlsx4 = kvutil.filename_unique(
    {
        "base_filename": "t_kvxlstest4",
        "file_ext": ".xlsx",
        "uniqtype": "datecnt",
        "overwrite": True,
        "forceuniq": True,
    }
)


xlsfloat2datetime = [
    43080.0,
    datetime.datetime.strptime("12/11/2017", "%m/%d/%Y"),
    "12/11/2017",
]

records = [
    {
        "Company": "NHLiq",
        "Wine": "Caravan Cabernet Sauvignon",
        "Vintage_Wine": "Caravan Cabernet Sauvignon 2014",
        "Vintage": "2014",
        "Date": "12/11/2017",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "BevMo",
        "Wine": "Caymus Cabernet Sauvignon Napa (750 ML)",
        "Vintage_Wine": "Caymus Cabernet Sauvignon Napa (750 ML) 2014",
        "Vintage": "2014",
        "Date": "10/31/2015",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "BevMo",
        "Wine": "Caymus Cabernet Sauvignon Special Select (750 ML)",
        "Vintage_Wine": "Caymus Cabernet Sauvignon Special Select (750 ML) 2014",
        "Vintage": "2014",
        "Date": "10/16/2016",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "BevMo",
        "Wine": "Caymus Cabernet Special Select (750 ML)",
        "Vintage_Wine": "Caymus Cabernet Special Select (750 ML) 2014",
        "Vintage": "2014",
        "Date": "7/23/2015",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "WineClub",
        "Wine": "CAYMUS VINEYARDS CABERNET SAUVIGNON",
        "Vintage_Wine": "CAYMUS VINEYARDS CABERNET SAUVIGNON 2014",
        "Vintage": "2014",
        "Date": "7/15/2018",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "WineClub",
        "Wine": "CAYMUS VINEYARDS SPECIAL SELECTION CABERNET SAUVIGNON",
        "Vintage_Wine": "CAYMUS VINEYARDS SPECIAL SELECTION CABERNET SAUVIGNON 2014",
        "Vintage": "2014",
        "Date": "7/15/2018",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "BevMo",
        "Wine": "Chappellet Cab Sauv Pritchard (750 ML)",
        "Vintage_Wine": "Chappellet Cab Sauv Pritchard (750 ML) 2014",
        "Vintage": "2014",
        "Date": "12/31/2015",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "TotalCA",
        "Wine": "Chappellet Cabernet Sauvignon Napa Signature",
        "Vintage_Wine": "Chappellet Cabernet Sauvignon Napa Signature 2014",
        "Vintage": "2014",
        "Date": "7/2/2013",
        "Type": "red-cab",
        "LastSeen": "2/27/2016",
    },
    {
        "Company": "WineClub",
        "Wine": "CHAPPELLET PRITCHARD HILL CABERNET SAUVIGNON",
        "Vintage_Wine": "CHAPPELLET PRITCHARD HILL CABERNET SAUVIGNON 2014",
        "Vintage": "2014",
        "Date": "7/15/2018",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "WineClub",
        "Wine": "CHAPPELLET SIGNATURE RESERVE CABERNET SAUVIGNON",
        "Vintage_Wine": "CHAPPELLET SIGNATURE RESERVE CABERNET SAUVIGNON 2014",
        "Vintage": "2014",
        "Date": "7/15/2018",
        "Type": "red-cab",
        "LastSeen": "Never",
    },
    {
        "Company": "Vons",
        "Wine": "Charles Krug Cabernet Sauvignon Wine - 750 Ml",
        "Vintage_Wine": "Charles Krug Cabernet Sauvignon Wine 2014 - 750 Ml",
        "Vintage": "2014",
        "Date": "2/21/2009",
        "Type": "red-cab",
        "LastSeen": "2/27/2016",
    },
]

records_multi_key = {
    "BevMo": {
        "Caymus Cabernet Sauvignon Napa (750 ML)": {"2014": 2},
        "Caymus Cabernet Sauvignon Special Select (750 ML)": {"2014": 3},
        "Caymus Cabernet Special Select (750 ML)": {"2014": 4},
        "Chappellet Cab Sauv Pritchard (750 ML)": {"2014": 7},
    },
    "NHLiq": {"Caravan Cabernet Sauvignon": {"2014": 1}},
    "TotalCA": {"Chappellet Cabernet Sauvignon Napa Signature": {"2014": 8}},
    "Vons": {"Charles Krug Cabernet Sauvignon Wine - 750 Ml": {"2014": 11}},
    "WineClub": {
        "CAYMUS VINEYARDS CABERNET SAUVIGNON": {"2014": 5},
        "CAYMUS VINEYARDS SPECIAL SELECTION CABERNET SAUVIGNON": {"2014": 6},
        "CHAPPELLET PRITCHARD HILL CABERNET SAUVIGNON": {"2014": 9},
        "CHAPPELLET SIGNATURE RESERVE CABERNET SAUVIGNON": {"2014": 10},
    },
}

records2 = [
    {
        "StringField": "this is a string line 1",
        "DateField": datetime.datetime(2020, 1, 1, 0, 0),
        "IntField": 1,
        "NumberField": 12.34,
    },
    {
        "StringField": "this is line 2",
        "DateField": datetime.datetime(2020, 2, 2, 0, 0),
        "IntField": 10,
        "NumberField": 1234.56,
    },
]


business_keys = [
    "NHLiq|Caravan Cabernet Sauvignon",
    "BevMo|Caymus Cabernet Sauvignon Napa (750 ML)",
    "BevMo|Caymus Cabernet Sauvignon Special Select (750 ML)",
    "BevMo|Caymus Cabernet Special Select (750 ML)",
    "WineClub|CAYMUS VINEYARDS CABERNET SAUVIGNON",
    "WineClub|CAYMUS VINEYARDS SPECIAL SELECTION CABERNET SAUVIGNON",
    "BevMo|Chappellet Cab Sauv Pritchard (750 ML)",
    "TotalCA|Chappellet Cabernet Sauvignon Napa Signature",
    "WineClub|CHAPPELLET PRITCHARD HILL CABERNET SAUVIGNON",
    "WineClub|CHAPPELLET SIGNATURE RESERVE CABERNET SAUVIGNON",
    "Vons|Charles Krug Cabernet Sauvignon Wine - 750 Ml",
]


req_cols = ["Company", "Wine"]
req_cols2 = ["StringField", "DateField"]

xlatdict = {"Company": "NewCompany", "Wine": "Winery"}
req_cols_xlat = ["NewCompany", "Winery"]

# multi sheet xlsx created for testing chgsheet
optiondict41 = {"sheet_name": "Sheet", "replace_sheet": True}
optiondict42 = {"sheet_name": "set_sheet_name2", "replace_sheet": True}
req_cols4_1 = req_cols
req_cols4_2 = req_cols2

# create XLS to be tested with, add sheets to the existing file - because we enabled "replace_sheet"
kvxls.writelist2xls(filenamexls4, records, optiondict=optiondict41, debug=False)
kvxls.writelist2xls(
    filenamexls4, records2, optiondict=optiondict42, debug=False
)

# create XLSX to be tested with, add sheets to the existing file - because we enabled "replace_sheet"
kvxls.writelist2xls(
    filenamexlsx4, records, optiondict=optiondict41, debug=False
)
kvxls.writelist2xls(
    filenamexlsx4, records2, optiondict=optiondict42, debug=False
)

# --------------------------------------------------------------------------------


class TestKVxls(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # create xls file
        kvxls.writelist2xls(filenamexls, records)
        # create xlsx file
        kvxls.writelist2xls(filenamexlsx, records)
        # create xls file
        kvxls.writelist2xls(filenamexls2, records2)
        # create xlsx file
        kvxls.writelist2xls(filenamexlsx2, records2)

    @classmethod
    def tearDownClass(cls):
        if 1:
            kvutil.remove_filename(filenamexls, kvutil.functionName())
            kvutil.remove_filename(filenamexlsx, kvutil.functionName())
            kvutil.remove_filename(filenamexls2, kvutil.functionName())
            kvutil.remove_filename(filenamexlsx2, kvutil.functionName())
            kvutil.remove_filename(filenamexls4, kvutil.functionName())
            kvutil.remove_filename(filenamexlsx4, kvutil.functionName())

    # the function name: def create_excel_config(
    def test_create_excel_config_p01_pass(self):
        excel_cfg = kvxls.create_excel_config()
        self.assertTrue(isinstance(excel_cfg, kvxls.ExcelConfig))
        self.assertEqual(excel_cfg.xlsxfiletype, None)
        self.assertEqual(excel_cfg.filename, None)
        self.assertEqual(excel_cfg.req_cols, None)
        self.assertEqual(excel_cfg.col_aref, None)
        self.assertEqual(excel_cfg.xlatdict, None)
        self.assertFalse(excel_cfg.allow_empty)
        self.assertFalse(excel_cfg.aref_result)
        self.assertFalse(excel_cfg.col_header)
        self.assertTrue(excel_cfg.data_only)
        self.assertTrue(excel_cfg.keep_vba)
        self.assertFalse(excel_cfg.no_header)
        self.assertFalse(excel_cfg.no_warnings)
        self.assertEqual(excel_cfg.start_row, 0)
        self.assertEqual(excel_cfg.max_rows, 100000000)
        self.assertEqual(excel_cfg.row_header, None)
        self.assertFalse(excel_cfg.save_row)
        self.assertFalse(excel_cfg.save_row_abs)
        self.assertFalse(excel_cfg.save_col_abs)
        self.assertFalse(excel_cfg.save_col_fmt)
        self.assertFalse(excel_cfg.replace_sheet)
        self.assertEqual(excel_cfg.replace_index, None)
    def test_create_excel_config_p02_optiondict(self):
        optiondict = {
            'save_row': True,
            'col_header': True,
        }
        saved_optiondict = optiondict.copy()
        excel_cfg = kvxls.create_excel_config(optiondict)
        self.assertTrue(isinstance(excel_cfg, kvxls.ExcelConfig))
        self.assertEqual(excel_cfg.xlsxfiletype, None)
        self.assertEqual(excel_cfg.filename, None)
        self.assertEqual(excel_cfg.req_cols, None)
        self.assertEqual(excel_cfg.col_aref, None)
        self.assertEqual(excel_cfg.xlatdict, None)
        self.assertFalse(excel_cfg.allow_empty)
        self.assertFalse(excel_cfg.aref_result)
        self.assertTrue(excel_cfg.col_header)
        self.assertTrue(excel_cfg.data_only)
        self.assertTrue(excel_cfg.keep_vba)
        self.assertFalse(excel_cfg.no_header)
        self.assertFalse(excel_cfg.no_warnings)
        self.assertEqual(excel_cfg.start_row, 0)
        self.assertEqual(excel_cfg.max_rows, 100000000)
        self.assertEqual(excel_cfg.row_header, None)
        self.assertTrue(excel_cfg.save_row)
        self.assertFalse(excel_cfg.save_row_abs)
        self.assertFalse(excel_cfg.save_col_abs)
        self.assertFalse(excel_cfg.save_col_fmt)
        self.assertFalse(excel_cfg.replace_sheet)
        self.assertEqual(excel_cfg.replace_index, None)

        self.assertEqual(optiondict, saved_optiondict)
    def test_create_excel_config_p03_badoptiondict(self):
        saved_optiondict = {
            'save_row': True,
            'col_header': True,
            'saverow': True,
            'colheader': True,
        }
        optiondict = {
            'saverow': True,
            'colheader': True,
        }

        excel_cfg = kvxls.create_excel_config(optiondict)
        self.assertTrue(isinstance(excel_cfg, kvxls.ExcelConfig))
        self.assertEqual(excel_cfg.xlsxfiletype, None)
        self.assertEqual(excel_cfg.filename, None)
        self.assertEqual(excel_cfg.req_cols, None)
        self.assertEqual(excel_cfg.col_aref, None)
        self.assertEqual(excel_cfg.xlatdict, None)
        self.assertFalse(excel_cfg.allow_empty)
        self.assertFalse(excel_cfg.aref_result)
        self.assertTrue(excel_cfg.col_header)
        self.assertTrue(excel_cfg.data_only)
        self.assertTrue(excel_cfg.keep_vba)
        self.assertFalse(excel_cfg.no_header)
        self.assertFalse(excel_cfg.no_warnings)
        self.assertEqual(excel_cfg.start_row, 0)
        self.assertEqual(excel_cfg.max_rows, 100000000)
        self.assertEqual(excel_cfg.row_header, None)
        self.assertTrue(excel_cfg.save_row)
        self.assertFalse(excel_cfg.save_row_abs)
        self.assertFalse(excel_cfg.save_col_abs)
        self.assertFalse(excel_cfg.save_col_fmt)
        self.assertFalse(excel_cfg.replace_sheet)
        self.assertEqual(excel_cfg.replace_index, None)

        self.assertEqual(optiondict, saved_optiondict)
        
        
    # strip_xls_illegal_chars
    def test_strip_xls_illegal_chars_p01_chars_removed(self):
        newvalue = kvxls.strip_xls_illegal_chars(kvxls.ILLEGAL_CHARACTERS_STR)
        self.assertEqual(newvalue, " " * len(kvxls.ILLEGAL_CHARACTERS_STR))

    def test_strip_xls_illegal_chars_p02_chars_not_removed(self):
        goodchars = "ABCDefg123"
        self.assertEqual(kvxls.strip_xls_illegal_chars(goodchars), goodchars)

    def test_strip_xls_illegal_chars_p03_int(self):
        goodchars = 5
        self.assertEqual(kvxls.strip_xls_illegal_chars(goodchars), goodchars)

    def test_strip_xls_illegal_chars_p04_float(self):
        goodchars = 5.5
        self.assertEqual(kvxls.strip_xls_illegal_chars(goodchars), goodchars)

    def test_strip_xls_illegal_chars_p04_datetime(self):
        goodchars = datetime.datetime(2026, 3, 23, 13, 59)
        self.assertEqual(kvxls.strip_xls_illegal_chars(goodchars), goodchars)

    # xldate_to_datetime: convert excel date fields to python
    def test_xldate_to_datetime_p01_float(self):
        self.assertEqual(
            kvxls.xldate_to_datetime(xlsfloat2datetime[0]), xlsfloat2datetime[1]
        )

    def test_xldate_to_datetime_p02_string(self):
        self.assertEqual(
            kvxls.xldate_to_datetime(xlsfloat2datetime[2]), xlsfloat2datetime[1]
        )

    def test_xldate_to_datetime_p03_blank_skipblank(self):
        self.assertEqual(kvxls.xldate_to_datetime("", True), "")

    def test_xldate_to_datetime_f01_blank(self):
        with self.assertRaises(Exception) as context:
            kvxls.xldate_to_datetime("")

    # the function name: def _extract_excel_row_into_list(xlsxfiletype, s, row, colstart, colmax, debug=False):
    # def test__extract_excel_row_into_list_p01_pass(self):
    def test__extract_excel_row_into_list_p01_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        row, c_row, c_col = kvxls._extract_excel_row_into_list(
            excel_dict["xlsxfiletype"],
            excel_dict["s"],
            1,
            excel_dict["sheetmincol"],
            excel_dict["sheetmaxcol"],
            debug=False,
        )
        expected_result = [
            "NHLiq",
            "Caravan Cabernet Sauvignon",
            "Caravan Cabernet Sauvignon 2014",
            "2014",
            "12/11/2017",
            "red-cab",
            "Never",
        ]
        self.assertEqual(expected_result, row)
        self.assertEqual(2, c_row)
        self.assertEqual(1, c_col)

    ########################################
    # the function name: def getExcelCellValue(excel_dict, row, col_name, debug=False):
    def test_getExcelCellValue_p01_xlsx_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "NHLiq")
        value = kvxls.getExcelCellValue(excel_dict, 3, "Company")
        self.assertEqual(value, "BevMo")

    def test_getExcelCellValue_p02_xlsx_start_row_3(self):
        kvxls.writelist2xls(
            filenamexlsx3, records, optiondict={"start_row": 3}, debug=False
        )
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx3, req_cols, debug=False
        )
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "NHLiq")
        value = kvxls.getExcelCellValue(excel_dict, 3, "Company")
        self.assertEqual(value, "BevMo")
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    ########################################
    # the function name: def setExcelCellValue(excel_dict, row, col_name, value, debug=False):
    def test_setExcelCellValue_p01_xlsx_top(self):
        fstarter = "tst1-"
        kvxls.writelist2xls(fstarter + filenamexlsx3, records, debug=False)
        excel_dict = kvxls.readxls_findheader(
            fstarter + filenamexlsx3, req_cols, debug=False
        )
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "NHLiq")
        kvxls.setExcelCellValue(excel_dict, 1, "Company", "KVValue")
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "KVValue")
        excel_dict["keep_vba"] = False
        kvxls.writexls(excel_dict, fstarter + filenamexlsx3, debug=False)
        excel_dict = kvxls.readxls_findheader(
            fstarter + filenamexlsx3, req_cols, debug=False
        )
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "KVValue")
        kvutil.remove_filename(
            fstarter + filenamexlsx3, kvutil.functionName(), debug=False
        )

    def test_setExcelCellValue_p02_xlsx_start_row_3(self):
        fstarter = "tst2-"
        kvxls.writelist2xls(
            fstarter + filenamexlsx3,
            records,
            optiondict={"start_row": 3},
            debug=False,
        )
        excel_dict = kvxls.readxls_findheader(
            fstarter + filenamexlsx3, req_cols, debug=False
        )
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "NHLiq")
        kvxls.setExcelCellValue(excel_dict, 1, "Company", "KVValue")
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "KVValue")
        excel_dict["keep_vba"] = False
        kvxls.writexls(excel_dict, fstarter + filenamexlsx3, debug=False)
        excel_dict = kvxls.readxls_findheader(
            fstarter + filenamexlsx3, req_cols, debug=False
        )
        value = kvxls.getExcelCellValue(excel_dict, 1, "Company")
        self.assertEqual(value, "KVValue")
        kvutil.remove_filename(
            fstarter + filenamexlsx3, kvutil.functionName(), debug=False
        )

    # getExcelCellFont
    def test_getExcelCellFont_p01_xlsx_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        (name, size, bold, italic, underline, strike, color) = (
            kvxls.getExcelCellFont(
                excel_dict,
                1,
                "Company",
                debug=False,
            )
        )
        # print(name, size, bold, italic, underline, strike, color)
        self.assertEqual(name, "Calibri")
        self.assertEqual(size, 11.0)
        self.assertEqual(bold, False)
        self.assertEqual(italic, False)
        self.assertEqual(underline, None)
        self.assertEqual(strike, None)
        self.assertEqual(type(color), openpyxl.styles.colors.Color)

    # setExcelCellFont
    def test_setExcelCellFont_p01_xlsx_name_size(self):
        debug = False
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        (
            name_orig,
            size_orig,
            bold_orig,
            italic_orig,
            underline_orig,
            strike_orig,
            color_orig,
        ) = kvxls.getExcelCellFont(
            excel_dict,
            1,
            "Company",
            debug=False,
        )
        if debug:
            print("\norig values")
            print(
                name_orig,
                size_orig,
                bold_orig,
                italic_orig,
                underline_orig,
                strike_orig,
                color_orig,
            )
        # change
        kvxls.setExcelCellFont(
            excel_dict,
            1,
            "Company",
            name="Arial",
            size=14.0,
            debug=False,
        )
        (name, size, bold, italic, underline, strike, color) = (
            kvxls.getExcelCellFont(
                excel_dict,
                1,
                "Company",
                debug=False,
            )
        )
        if debug:
            print("\nchanged value")
            print(name, size, bold, italic, underline, strike, color)
        # return
        kvxls.setExcelCellFont(
            excel_dict,
            1,
            "Company",
            name=name_orig,
            size=size_orig,
            bold=bold_orig,
            italic=italic_orig,
            underline=underline_orig,
            strike=strike_orig,
            color=color_orig,
            debug=False,
        )
        self.assertEqual(name, "Arial")
        self.assertEqual(size, 14.0)
        self.assertEqual(bold, False)
        self.assertEqual(italic, False)
        self.assertEqual(underline, None)
        self.assertEqual(strike, None)
        self.assertEqual(type(color), openpyxl.styles.colors.Color)

    def test_setExcelCellFont_p02_xlsx_bold_italic_strike(self):
        debug = False
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        (
            name_orig,
            size_orig,
            bold_orig,
            italic_orig,
            underline_orig,
            strike_orig,
            color_orig,
        ) = kvxls.getExcelCellFont(
            excel_dict,
            1,
            "Company",
            debug=False,
        )
        if debug:
            print("\norig values")
            print(
                name_orig,
                size_orig,
                bold_orig,
                italic_orig,
                underline_orig,
                strike_orig,
                color_orig,
            )
        # change
        kvxls.setExcelCellFont(
            excel_dict,
            1,
            "Company",
            bold=True,
            italic=True,
            #            underline=True,
            strike=True,
            debug=False,
        )
        (name, size, bold, italic, underline, strike, color) = (
            kvxls.getExcelCellFont(
                excel_dict,
                1,
                "Company",
                debug=False,
            )
        )
        if debug:
            print("\nchanged value")
            print(name, size, bold, italic, underline, strike, color)
        # return
        kvxls.setExcelCellFont(
            excel_dict,
            1,
            "Company",
            name=name_orig,
            size=size_orig,
            bold=bold_orig,
            italic=italic_orig,
            underline=underline_orig,
            strike=strike_orig,
            color=color_orig,
            debug=False,
        )
        if debug:
            print(name, size, bold, italic, underline, strike, color)
            self.assertEqual(bold, True)
        self.assertEqual(italic, True)
        #        self.assertEqual(underline, True)
        self.assertEqual(strike, True)
        self.assertEqual(type(color), openpyxl.styles.colors.Color)

    def test_setExcelCellFont_p03_xlsx_underline(self):
        debug = False
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        (
            name_orig,
            size_orig,
            bold_orig,
            italic_orig,
            underline_orig,
            strike_orig,
            color_orig,
        ) = kvxls.getExcelCellFont(
            excel_dict,
            1,
            "Company",
            debug=False,
        )
        if debug:
            print("\norig values")
            print(
                name_orig,
                size_orig,
                bold_orig,
                italic_orig,
                underline_orig,
                strike_orig,
                color_orig,
            )
        # change
        kvxls.setExcelCellFont(
            excel_dict, 1, "Company", underline="single", debug=False
        )
        (name, size, bold, italic, underline, strike, color) = (
            kvxls.getExcelCellFont(
                excel_dict,
                1,
                "Company",
                debug=False,
            )
        )
        if debug:
            print("\nchanged value")
            print(name, size, bold, italic, underline, strike, color)
        # return
        kvxls.setExcelCellFont(
            excel_dict,
            1,
            "Company",
            name=name_orig,
            size=size_orig,
            bold=bold_orig,
            italic=italic_orig,
            underline=underline_orig,
            strike=strike_orig,
            color=color_orig,
            debug=False,
        )
        if debug:
            print(name, size, bold, italic, underline, strike, color)
            self.assertEqual(underline, "single")
        self.assertEqual(type(color), openpyxl.styles.colors.Color)

    def test_setExcelCellFont_p04_xlsx_color(self):
        debug = False
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        (
            name_orig,
            size_orig,
            bold_orig,
            italic_orig,
            underline_orig,
            strike_orig,
            color_orig,
        ) = kvxls.getExcelCellFont(
            excel_dict,
            1,
            "Company",
            debug=False,
        )
        if debug:
            print("\norig values")
            print(
                name_orig,
                size_orig,
                bold_orig,
                italic_orig,
                underline_orig,
                strike_orig,
                color_orig,
            )
        # change
        kvxls.setExcelCellFont(
            excel_dict, 1, "Company", color="FF0000", debug=False
        )
        (name, size, bold, italic, underline, strike, color) = (
            kvxls.getExcelCellFont(
                excel_dict,
                1,
                "Company",
                debug=False,
            )
        )
        if debug:
            print("\nchanged value")
            print(name, size, bold, italic, underline, strike, color)
        # return
        kvxls.setExcelCellFont(
            excel_dict,
            1,
            "Company",
            name=name_orig,
            size=size_orig,
            bold=bold_orig,
            italic=italic_orig,
            underline=underline_orig,
            strike=strike_orig,
            color=color_orig,
            debug=False,
        )
        if debug:
            print(name, size, bold, italic, underline, strike, color)
        self.assertEqual(color.rgb, "00FF0000")
        self.assertEqual(type(color), openpyxl.styles.colors.Color)

    ########################################
    # the function name: def getExcelCellPatternFill(excel_dict, row, col_name, debug=False):
    def test_getExcelCellPatternFill_p01_xlsx_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )

        cell_color, cell_fill_type, cell_start_color, cell_end_color = (
            kvxls.getExcelCellPatternFill(excel_dict, 1, "Company", debug=False)
        )
        self.assertEqual(cell_color, None)
        self.assertEqual(cell_fill_type, None)
        self.assertEqual(cell_start_color, None)
        self.assertEqual(cell_end_color, None)

        # print('CellPattern: ', cell_color, cell_fill_type, cell_start_color, cell_end_color)

    ########################################
    # the function name: def setExcelCellPatternFill(excel_dict, row, col_name, fill=None, start_color=None, end_color=None, fg_color=None, fill_type="solid", debug=False):
    def test_setExcelCellPatternFill_p01_xlsx_fg_color(self):
        debug = False
        yellow_fill = "FFFFFF00"
        row = 2
        fstarter = "tst3-"
        col_name = "Date"
        # open the file
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        excel_dict["keep_vba"] = False
        # save original values
        (
            orig_cell_color,
            orig_cell_fill_type,
            orig_cell_start_color,
            orig_cell_end_color,
        ) = kvxls.getExcelCellPatternFill(
            excel_dict, row, col_name, debug=False
        )
        if debug:
            print(
                "Orig Pattern1: ",
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
            print(
                [
                    type(x)
                    for x in (
                        orig_cell_color,
                        orig_cell_fill_type,
                        orig_cell_start_color,
                        orig_cell_end_color,
                    )
                ]
            )
        # update values
        (
            orig_cell_color,
            orig_cell_fill_type,
            orig_cell_start_color,
            orig_cell_end_color,
        ) = [
            "None" if x is None else x
            for x in (
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
        ]
        if debug:
            print(
                "Orig Pattern2: ",
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
            print(
                [
                    type(x)
                    for x in (
                        orig_cell_color,
                        orig_cell_fill_type,
                        orig_cell_start_color,
                        orig_cell_end_color,
                    )
                ]
            )
        # change
        kvxls.setExcelCellPatternFill(
            excel_dict, row, col_name, fg_color=yellow_fill
        )
        cell_color, cell_fill_type, cell_start_color, cell_end_color = (
            kvxls.getExcelCellPatternFill(
                excel_dict, row, col_name, debug=False
            )
        )
        if debug:
            print(
                "Updated Pattern: ",
                cell_color,
                cell_fill_type,
                cell_start_color,
                cell_end_color,
            )
            self.assertEqual(yellow_fill, cell_color)
        self.assertEqual("solid", cell_fill_type)
        # restor
        kvxls.setExcelCellPatternFill(
            excel_dict,
            row,
            col_name,
            fg_color=orig_cell_color,
            fill_type=orig_cell_fill_type,
            start_color=orig_cell_start_color,
            end_color=orig_cell_end_color,
        )
        cell_color, cell_fill_type, cell_start_color, cell_end_color = (
            kvxls.getExcelCellPatternFill(
                excel_dict, row, col_name, debug=False
            )
        )
        if debug:
            print(
                "Restored Pattern: ",
                cell_color,
                cell_fill_type,
                cell_start_color,
                cell_end_color,
            )

    def test_setExcelCellPatternFill_p02_xlsx_fill(self):
        debug = False
        start_color = "00FF0000"  # red
        end_color = "00FF0000"  # red
        fill_type = "solid"
        row = 2
        fstarter = "tst4-"
        col_name = "Date"
        # create the fill object
        red_fill = PatternFill(
            start_color=start_color, end_color=end_color, fill_type=fill_type
        )
        # open the file
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        excel_dict["keep_vba"] = False
        # capture the orginal values
        (
            orig_cell_color,
            orig_cell_fill_type,
            orig_cell_start_color,
            orig_cell_end_color,
        ) = kvxls.getExcelCellPatternFill(
            excel_dict, row, col_name, debug=False
        )
        if debug:
            print(
                "Orig Pattern1: ",
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
            print(
                [
                    type(x)
                    for x in (
                        orig_cell_color,
                        orig_cell_fill_type,
                        orig_cell_start_color,
                        orig_cell_end_color,
                    )
                ]
            )
        # update values - to make None a string 'None'
        (
            orig_cell_color,
            orig_cell_fill_type,
            orig_cell_start_color,
            orig_cell_end_color,
        ) = [
            "None" if x is None else x
            for x in (
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
        ]
        if debug:
            print(
                "Orig Pattern2: ",
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
            print(
                [
                    type(x)
                    for x in (
                        orig_cell_color,
                        orig_cell_fill_type,
                        orig_cell_start_color,
                        orig_cell_end_color,
                    )
                ]
            )
        # change the pattern
        kvxls.setExcelCellPatternFill(
            excel_dict, row, col_name, fill=red_fill, debug=False
        )
        # read in after the change
        cell_color, cell_fill_type, cell_start_color, cell_end_color = (
            kvxls.getExcelCellPatternFill(
                excel_dict, row, col_name, debug=False
            )
        )
        if debug:
            print(
                "Updated Pattern: ",
                cell_color,
                cell_fill_type,
                cell_start_color,
                cell_end_color,
            )
        # test this
        self.assertEqual(start_color, cell_color)
        self.assertEqual(fill_type, cell_fill_type)
        self.assertEqual(cell_start_color.rgb, start_color)
        self.assertEqual(cell_end_color.rgb, end_color)
        # restor
        kvxls.setExcelCellPatternFill(
            excel_dict,
            row,
            col_name,
            fg_color=orig_cell_color,
            fill_type=orig_cell_fill_type,
            start_color=orig_cell_start_color,
            end_color=orig_cell_end_color,
            debug=False,
        )
        cell_color, cell_fill_type, cell_start_color, cell_end_color = (
            kvxls.getExcelCellPatternFill(
                excel_dict, row, col_name, debug=False
            )
        )
        if debug:
            print(
                "Restored Pattern: ",
                cell_color,
                cell_fill_type,
                cell_start_color,
                cell_end_color,
            )

    def test_setExcelCellPatternFill_p03_xlsx_start_end_color(self):
        debug = False
        start_color = "00FF0000"  # red
        end_color = "00FF0000"  # red
        yellow_fill = "FFFFFF00"
        fill_type = "solid"
        row = 2
        fstarter = "tst5-"
        col_name = "Date"
        # open the file
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        excel_dict["keep_vba"] = False
        # save original values
        (
            orig_cell_color,
            orig_cell_fill_type,
            orig_cell_start_color,
            orig_cell_end_color,
        ) = kvxls.getExcelCellPatternFill(
            excel_dict, row, col_name, debug=False
        )
        if debug:
            print(
                "Orig Pattern1: ",
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
            print(
                [
                    type(x)
                    for x in (
                        orig_cell_color,
                        orig_cell_fill_type,
                        orig_cell_start_color,
                        orig_cell_end_color,
                    )
                ]
            )
        # update values
        (
            orig_cell_color,
            orig_cell_fill_type,
            orig_cell_start_color,
            orig_cell_end_color,
        ) = [
            "None" if x is None else x
            for x in (
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
        ]
        if debug:
            print(
                "Orig Pattern2: ",
                orig_cell_color,
                orig_cell_fill_type,
                orig_cell_start_color,
                orig_cell_end_color,
            )
            print(
                [
                    type(x)
                    for x in (
                        orig_cell_color,
                        orig_cell_fill_type,
                        orig_cell_start_color,
                        orig_cell_end_color,
                    )
                ]
            )
        # change
        kvxls.setExcelCellPatternFill(
            excel_dict,
            row,
            col_name,
            start_color=start_color,
            end_color=end_color,
            debug=False,
        )
        cell_color, cell_fill_type, cell_start_color, cell_end_color = (
            kvxls.getExcelCellPatternFill(
                excel_dict, row, col_name, debug=False
            )
        )
        if debug:
            print(
                "Updated Pattern: ",
                cell_color,
                cell_fill_type,
                cell_start_color,
                cell_end_color,
            )
            self.assertEqual(yellow_fill, cell_color)
        # tests
        self.assertEqual(start_color, cell_color)
        self.assertEqual(fill_type, cell_fill_type)
        self.assertEqual(cell_start_color.rgb, start_color)
        self.assertEqual(cell_end_color.rgb, end_color)
        # restor
        kvxls.setExcelCellPatternFill(
            excel_dict,
            row,
            col_name,
            fg_color=orig_cell_color,
            fill_type=orig_cell_fill_type,
            start_color=orig_cell_start_color,
            end_color=orig_cell_end_color,
        )
        cell_color, cell_fill_type, cell_start_color, cell_end_color = (
            kvxls.getExcelCellPatternFill(
                excel_dict, row, col_name, debug=False
            )
        )
        if debug:
            print(
                "Restored Pattern: ",
                cell_color,
                cell_fill_type,
                cell_start_color,
                cell_end_color,
            )

    ########################################
    # the function name: def copyExcelCellFmtOnRow(excel_dict_src, src_row, excel_dict_out, row, debug=False):
    def test_copyExcelCellFmtOnRow_p01_pass(self):
        yellow = "FFFFFF00"
        # open the file
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        # set the formatting on this file
        for colname in excel_dict['header']:
            kvxls.setExcelCellPatternFill(excel_dict, 2, colname, fg_color=yellow, fill_type='solid', debug=False)
        # create a file to copy over to
        fstarter = "tst1-"
        new_file = fstarter + filenamexlsx3
        kvxls.writelist2xls(new_file, records, debug=False)
        excel_dict2 = kvxls.readxls_findheader(
            new_file, req_cols, debug=False
        )
        # now copy over the format
        kvxls.copyExcelCellFmtOnRow(excel_dict, 2, excel_dict2, 2, debug=False)
        # test that this worked
        for colname in excel_dict2['header']:
            cell_color, cell_fill_type, cell_start_color, cell_end_color = kvxls.getExcelCellPatternFill(excel_dict2, 2, colname)
            self.assertEqual(cell_color, yellow)
        # remove the temp file
        kvutil.remove_filename(new_file)


    # the function name: def setExcelColumnValue(excel_dict, col_name, value='', debug=False):
    def test_setExcelColumnValue_p01_pass(self):
        col_name = "LastSeen"
        value = "SetValue"
        # open the file
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        excel_dict["keep_vba"] = False
        # now populated the column
        kvxls.setExcelColumnValue(
            excel_dict=excel_dict, col_name=col_name, value=value, debug=False
        )
        # now pull values and check
        value_read = kvxls.getExcelCellValue(
            excel_dict,
            1,
            col_name,
        )
        self.assertEqual(value_read, value)
        # now pull values and check
        value_read = kvxls.getExcelCellValue(
            excel_dict,
            2,
            col_name,
        )
        self.assertEqual(value_read, value)

    ########################################
    # prior function: setExcelColumnValue
    # the function name: def any_field_is_populated(rec: dict, fldlist: list[str], debug: bool=False) -> bool:
    def test_any_field_is_populated_p01_pass(self):
        self.assertTrue(
            kvxls.any_field_is_populated(records[0], ["Company", "Wine"])
        )

    def test_any_field_is_populated_p02_no_values(self):
        self.assertFalse(
            kvxls.any_field_is_populated(
                {"Company": "", "Wine": ""}, ["Company", "Wine"]
            )
        )

    def test_any_field_is_populated_p03_no_value_not_string(self):
        self.assertTrue(
            kvxls.any_field_is_populated(
                {"Company": "", "Wine": 0}, ["Company", "Wine"]
            )
        )
        self.assertTrue(
            kvxls.any_field_is_populated(
                {"Company": "", "Wine": []}, ["Company", "Wine"]
            )
        )
        self.assertTrue(
            kvxls.any_field_is_populated(
                {"Company": "", "Wine": False}, ["Company", "Wine"]
            )
        )

    def test_any_field_is_populated_f01_rec_list_not_dict(self):
        with self.assertRaises(Exception) as context:
            kvxls.any_field_is_populated([1, 2], [])

    def test_any_field_is_populated_f01_fldlist_not_list(self):
        with self.assertRaises(Exception) as context:
            kvxls.any_field_is_populated(records[0], "string")

    ########################################
    # the function name: def create_multi_key_lookup_excel(excel_dict, fldlist, copy_fields=None):
    def test_create_multi_key_lookup_excel_p01_xlsx_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        multi_key_dict = kvxls.create_multi_key_lookup_excel(
            excel_dict, ["Company", "Wine", "Vintage"]
        )
        self.assertTrue(multi_key_dict, records_multi_key)

    def test_create_multi_key_lookup_excel_f01_xlsx_fldlist_dict(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols, debug=False
            )
            kvxls.test_create_multi_key_lookup_excel(
                excel_dict,
                {"a": 1},
                debug=False,
            )

    def test_create_multi_key_lookup_excel_f02_xlsx_fldlist_badvalues(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols, debug=False
            )
            kvxls.test_create_multi_key_lookup_excel(
                excel_dict,
                ["notkey", "notkey2"],
                debug=False,
            )

    def test_create_multi_key_lookup_excel_f03__xlsx_copy_fields_dict(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols, debug=False
            )
            kvxls.test_create_multi_key_lookup_excel(
                excel_dict,
                ["Company", "Wine", "Vintage"],
                copy_fields={"a": 1},
                debug=False,
            )

    def test_create_multi_key_lookup_excel_f04__xlsx__copy_fields_badvalues(
        self,
    ):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols, debug=False
            )
            kvxls.test_create_multi_key_lookup_excel(
                excel_dict,
                ["Company", "Wine", "Vintage"],
                copy_fields=["notkey", "notkey2"],
                debug=False,
            )

    ########################################
    # the function name: def calc_col_mapping(rec: dict) -> tuple[str, dict]:
    def test_calc_col_mapping_p01_pass(self):
        results = kvxls.readxls2list(
            filenamexlsx, optiondict={"save_col_abs": True}, debug=False
        )
        col_map_str, col_map = kvxls.calc_col_mapping(results[0])
        col_map_answer = '{"Company": 1, "Wine": 2, "Vintage_Wine": 3, "Vintage": 4, "Date": 5, "Type": 6, "LastSeen": 7, "XLSColAbs1": 8}'
        self.assertEqual(col_map_str, col_map_answer)

    def test_calc_col_mapping_f01_abs_col_not_used(self):
        results = kvxls.readxls2list(filenamexlsx, debug=False)
        with self.assertRaises(Exception) as context:
            col_map_str, col_map = kvxls.calc_col_mapping(results[0])

    ########################################
    # the function name: def set_col_mapping(rec) -> None:
    def test_set_col_mapping_p01_pass(self):
        results = kvxls.readxls2list(
            filenamexlsx, optiondict={"save_col_abs": True}, debug=False
        )
        kvxls.set_col_mapping(results[0])
        col_map_answer = '{"Company": 1, "Wine": 2, "Vintage_Wine": 3, "Vintage": 4, "Date": 5, "Type": 6, "LastSeen": 7, "XLSColAbs1": 8}'
        self.assertEqual(results[0][kvxls.FLD_XLSNEW_COLMAP], col_map_answer)

    def test_set_col_mapping_f01_abs_col_not_used(self):
        results = kvxls.readxls2list(filenamexlsx, debug=False)
        with self.assertRaises(Exception) as context:
            kvxls.set_col_mapping(results[0])

    ########################################
    # the function name: def set_col_mapping_list(records: list[dict]) -> None:
    def test_set_col_mapping_list_p01_pass(self):
        results = kvxls.readxls2list(
            filenamexlsx, optiondict={"save_col_abs": True}, debug=False
        )
        kvxls.set_col_mapping_list(results)
        col_map_answer = '{"Company": 1, "Wine": 2, "Vintage_Wine": 3, "Vintage": 4, "Date": 5, "Type": 6, "LastSeen": 7, "XLSColAbs1": 8}'
        self.assertEqual(results[0][kvxls.FLD_XLSNEW_COLMAP], col_map_answer)
        self.assertEqual(results[-1][kvxls.FLD_XLSNEW_COLMAP], col_map_answer)

    def test_set_col_mapping_list_f01_abs_col_not_used(self):
        results = kvxls.readxls2list(filenamexlsx, debug=False)
        with self.assertRaises(Exception) as context:
            kvxls.set_col_mapping_list(results)

    ########################################
    # the function name: def extract_col_mapping(rec: dict) -> tuple[dict, str]:
    def test_extract_col_mapping_p01_pass(self):
        results = kvxls.readxls2list(
            filenamexlsx, optiondict={"save_col_abs": True}, debug=False
        )
        kvxls.set_col_mapping(results[0])
        col_mapping, col_mapping_str = kvxls.extract_col_mapping(results[0])
        col_map_answer = '{"Company": 1, "Wine": 2, "Vintage_Wine": 3, "Vintage": 4, "Date": 5, "Type": 6, "LastSeen": 7, "XLSColAbs1": 8}'
        self.assertEqual(col_mapping_str, col_map_answer)

    def test_extract_col_mapping_f01_abs_col_not_used(self):
        results = kvxls.readxls2list(filenamexlsx, debug=False)
        with self.assertRaises(Exception) as context:
            kvxls.extract_col_mapping(results[0])

    ########################################
    # the function name: def readxls2list(xlsfile, sheetname=None, save_row=False, debug=False, optiondict=None):
    def test_readxls2list_p01_xls_pass(self):
        results = kvxls.readxls2list(filenamexls, debug=False)
        self.assertEqual(len(results), len(records))
        self.assertEqual(results, records)

    def test_readxls2list_p02_xlsx_pass(self):
        results = kvxls.readxls2list(filenamexlsx, debug=False)
        self.assertEqual(len(results), len(records))
        self.assertEqual(results, records)

    # save_row
    def test_readxls2list_p01_xls_save_row(self):
        results = kvxls.readxls2list(filenamexls, save_row=True, debug=False)
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertEqual(len(results), len(records))

    def test_readxls2list_p02_xlsx_save_row(self):
        results = kvxls.readxls2list(filenamexlsx, save_row=True, debug=False)
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertEqual(len(results), len(records))

    def test_readxls2list_p03_xls_save_row_abs(self):
        optiondict = {"save_row_abs": True}
        results = kvxls.readxls2list(
            filenamexls, save_row=True, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSRowAbs",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSRowAbs" in results[0].keys())
        self.assertEqual(len(results), len(records))

    def test_readxls2list_p03_xlsx_save_row_abs(self):
        optiondict = {"save_row_abs": True}
        results = kvxls.readxls2list(
            filenamexlsx, save_row=True, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSRowAbs",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSRowAbs" in results[0].keys())
        self.assertEqual(len(results), len(records))

    def test_readxls2list_p05_xls_save_col_abs(self):
        optiondict = {"save_col_abs": True}
        results = kvxls.readxls2list(
            filenamexls, save_row=True, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSColAbs1" in results[0].keys())
        self.assertEqual(len(results), len(records))

    def test_readxls2list_p06_xlsx_save_col_abs(self):
        optiondict = {"save_col_abs": True}
        results = kvxls.readxls2list(
            filenamexlsx, save_row=True, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSColAbs1" in results[0].keys())
        self.assertEqual(len(results), len(records))

    def test_readxls2list_p07_xls_save_row_and_col_abs(self):
        optiondict = {"save_col_abs": True, "save_row_abs": True}
        results = kvxls.readxls2list(
            filenamexls, save_row=True, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSRowAbs",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSRowAbs" in results[0].keys())
        self.assertTrue("XLSColAbs1" in results[0].keys())
        self.assertEqual(len(results), len(records))

    def test_readxls2list_p08_xlsx_save_row_and_col_abs(self):
        optiondict = {"save_col_abs": True, "save_row_abs": True}
        results = kvxls.readxls2list(
            filenamexlsx, save_row=True, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSRowAbs",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSRowAbs" in results[0].keys())
        self.assertTrue("XLSColAbs1" in results[0].keys())
        self.assertEqual(len(results), len(records))

    # sheet_name and save_row
    def test_readxls2list_p01_xls_sheet_name_save_row(self):
        results = kvxls.readxls2list(
            filenamexls4,
            sheetname=optiondict42["sheet_name"],
            save_row=True,
            debug=False,
        )
        self.assertEqual(
            list(results[0].keys()),
            ["StringField", "DateField", "IntField", "NumberField", "XLSRow"],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertEqual(len(results), len(records2))

    def test_readxls2list_p02_xlsx_sheet_name_save_row(self):
        results = kvxls.readxls2list(
            filenamexlsx4,
            sheetname=optiondict42["sheet_name"],
            save_row=True,
            debug=False,
        )
        self.assertEqual(
            list(results[0].keys()),
            ["StringField", "DateField", "IntField", "NumberField", "XLSRow"],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertEqual(len(results), len(records2))

    def test_readxls2list_p03_xls_sheet_name_save_row_and_abs(self):

        optiondict = {"save_col_abs": True, "save_row_abs": True}
        results = kvxls.readxls2list(
            filenamexls4,
            sheetname=optiondict42["sheet_name"],
            save_row=True,
            optiondict=optiondict,
            debug=False,
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "StringField",
                "DateField",
                "IntField",
                "NumberField",
                "XLSRow",
                "XLSRowAbs",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSRowAbs" in results[0].keys())
        self.assertTrue("XLSColAbs1" in results[0].keys())
        self.assertEqual(len(results), len(records2))

    def test_readxls2list_p04_xlsx_sheet_name_save_row_and_abs(self):
        optiondict = {"save_col_abs": True, "save_row_abs": True}
        results = kvxls.readxls2list(
            filenamexlsx4,
            sheetname=optiondict42["sheet_name"],
            save_row=True,
            optiondict=optiondict,
            debug=False,
        )
        self.assertEqual(
            list(results[0].keys()),
            [
                "StringField",
                "DateField",
                "IntField",
                "NumberField",
                "XLSRow",
                "XLSRowAbs",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results[0].keys())
        self.assertTrue("XLSRowAbs" in results[0].keys())
        self.assertTrue("XLSColAbs1" in results[0].keys())
        self.assertEqual(len(results), len(records2))

    def test_readxls2list_p01_xls_start_row(self):
        # set start_row
        kvxls.writelist2xls(
            filenamexls3, records, optiondict={"start_row": 3}, debug=False
        )
        results2 = kvxls.readxls2list_findheader(
            filenamexls3,
            req_cols=req_cols,
            optiondict={"save_row": True},
            debug=False,
        )
        self.assertEqual(
            list(results2[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
            ],
        )
        self.assertTrue("XLSRow" in results2[0].keys())
        self.assertEqual(len(results2), len(records))
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_p02_xlsx_start_row(self):
        kvxls.writelist2xls(
            filenamexlsx3, records, optiondict={"start_row": 3}, debug=False
        )
        results2 = kvxls.readxls2list_findheader(
            filenamexlsx3,
            req_cols=req_cols,
            optiondict={"save_row": True},
            debug=False,
        )
        self.assertEqual(
            list(results2[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
            ],
        )
        self.assertTrue("XLSRow" in results2[0].keys())
        self.assertEqual(len(results2), len(records))
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_readxls2list_p03_xls_start_row_and_abs(self):
        # set start_row with abs row/column
        optiondict = {
            "save_row": True,
            "save_col_abs": True,
            "save_row_abs": True,
        }
        kvxls.writelist2xls(
            filenamexls3, records, optiondict={"start_row": 3}, debug=False
        )
        results2 = kvxls.readxls2list_findheader(
            filenamexls3, req_cols=req_cols, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results2[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSRowAbs",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results2[0].keys())
        self.assertTrue("XLSRowAbs" in results2[0].keys())
        self.assertTrue("XLSColAbs1" in results2[0].keys())
        self.assertEqual(len(results2), len(records))
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_p04_xlsx_start_row_and_abs(self):
        optiondict = {
            "save_row": True,
            "save_col_abs": True,
            "save_row_abs": True,
            "start_row": 3,
        }
        kvxls.writelist2xls(
            filenamexlsx3, records, optiondict={"start_row": 3}, debug=False
        )
        results2 = kvxls.readxls2list_findheader(
            filenamexlsx3, req_cols=req_cols, optiondict=optiondict, debug=False
        )
        self.assertEqual(
            list(results2[0].keys()),
            [
                "Company",
                "Wine",
                "Vintage_Wine",
                "Vintage",
                "Date",
                "Type",
                "LastSeen",
                "XLSRow",
                "XLSRowAbs",
                "XLSColAbs1",
            ],
        )
        self.assertTrue("XLSRow" in results2[0].keys())
        self.assertTrue("XLSRowAbs" in results2[0].keys())
        self.assertTrue("XLSColAbs1" in results2[0].keys())
        self.assertEqual(len(results2), len(records))
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    ########################################
    # the function name: def readxls2dict(xlsfile, dictkeys, sheetname=None, save_row=False, dupkeyfail=False, debug=False, optiondict=None):
    def test_readxls2dict_p01_xls_pass(self):
        results = kvxls.readxls2dict(filenamexls, req_cols, debug=False)
        self.assertEqual(list(results.keys()), business_keys)

    def test_readxls2dict_p02_xlsx_pass(self):
        results = kvxls.readxls2dict(filenamexlsx, req_cols, debug=False)
        self.assertEqual(list(results.keys()), business_keys)

    # the function name: def readxls2dump(xlsfile, rows=10, sep=':', no_warnings=False, returnrecs=False, sheet_name_col=None, debug=False):
    def test_readxls2dump_p01_xls_pass(self):
        result = kvxls.readxls2dump(filenamexls, debug=False)
        expected_result = [
            "xlsfile:sheet_name:reccnt:colcnt:value:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:000:Company:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:001:Wine:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:002:Vintage_Wine:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:003:Vintage:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:004:Date:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:005:Type:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:006:LastSeen:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:007:1:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:000:NHLiq:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:001:Caravan Cabernet Sauvignon:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:002:Caravan Cabernet Sauvignon 2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:003:2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:004:12/11/2017:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:005:red-cab:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:006:Never:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:007:2:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:000:BevMo:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:001:Caymus Cabernet Sauvignon Napa (750 ML):",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:002:Caymus Cabernet Sauvignon Napa (750 ML) 2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:003:2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:004:10/31/2015:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:005:red-cab:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:006:Never:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:007:3:",
        ]
        run_result = []
        for x in expected_result:
            run_result.append(
                x.replace("t_kvxlstest-20250111v01.xls", filenamexls)
            )
        self.assertEqual(result[: len(run_result)], run_result)

    def test_readxls2dump_p02_xlsx_pass(self):
        result = kvxls.readxls2dump(filenamexlsx, debug=False)
        expected_result = [
            "xlsfile:sheet_name:reccnt:colcnt:value:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:000:Company:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:001:Wine:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:002:Vintage_Wine:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:003:Vintage:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:004:Date:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:005:Type:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:006:LastSeen:",
            "t_kvxlstest-20250111v01.xls:Sheet1:00:007:1:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:000:NHLiq:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:001:Caravan Cabernet Sauvignon:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:002:Caravan Cabernet Sauvignon 2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:003:2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:004:12/11/2017:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:005:red-cab:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:006:Never:",
            "t_kvxlstest-20250111v01.xls:Sheet1:01:007:2:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:000:BevMo:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:001:Caymus Cabernet Sauvignon Napa (750 ML):",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:002:Caymus Cabernet Sauvignon Napa (750 ML) 2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:003:2014:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:004:10/31/2015:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:005:red-cab:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:006:Never:",
            "t_kvxlstest-20250111v01.xls:Sheet1:02:007:3:",
        ]
        run_result = []
        for x in expected_result:
            run_result.append(
                x.replace("t_kvxlstest-20250111v01.xls", filenamexlsx).replace(
                    "Sheet1", "Sheet"
                )
            )
        self.assertEqual(result[: len(run_result)], run_result)

    ########################################
    # prior function: readxls2dump
    # the function name: def readxls2list_all_sheets(
    def test_readxls2list_all_sheets_p01_pass(self):
        # create sheet first time - so force a rebuild
        filename = kvxls.writelist2xls(filenamexlsx3, records, debug=False)
        # create optoins to add sheets
        optiondict = {"sheetname": "sheet2", "replace_sheet": True}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict, debug=False
        )
        # create another sheet
        optiondict = {"sheetname": "sheet3", "replace_sheet": True}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict, debug=False
        )
        # now read it in
        results, header, by_sheetname = kvxls.readxls2list_all_sheets(
            filenamexlsx3, req_cols
        )
        # tests
        self.assertEqual(list(results.keys()), ["Sheet", "sheet2", "sheet3"])
        self.assertEqual(results["Sheet"], records)
        self.assertFalse(by_sheetname)
        self.assertEqual(header, list(records[0].keys()))
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(
            filenamexlsx3, kvutil.functionName(), debug=False
        )

    def test_readxls2list_all_sheets_p02_blank_sheet(self):
        # create sheet first time - so force a rebuild
        filename = kvxls.writelist2xls(filenamexlsx3, records, debug=False)
        # create optoins to add sheets
        optiondict = {"sheetname": "sheet2", "replace_sheet": True}
        filename = kvxls.writelist2xls(
            filenamexlsx3, [], optiondict=optiondict, debug=False
        )
        # now read it in
        results, header, by_sheetname = kvxls.readxls2list_all_sheets(
            filenamexlsx3, req_cols
        )
        # error message
        errmsg = {
            "sheet2": "Max search row count [1] exceeded at row [1] - no match found"
        }
        # tests
        self.assertEqual(list(results.keys()), ["Sheet", "sheet2"])
        self.assertEqual(results["Sheet"], records)
        self.assertEqual(by_sheetname, errmsg)
        self.assertEqual(header, list(records[0].keys()))
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(
            filenamexlsx3, kvutil.functionName(), debug=False
        )

    ########################################
    # the function name: def readxls_excelDict(
    # def test_readxls_excelDict_p01_pass(self):
    ########################################

    ########################################
    # prior function: readxls_findheader
    # the function name: def readxls2excel_dict_findheader(
    # def test_readxls2excel_dict_findheader_p01_pass(self):
    ########################################

    ########################################
    # prior function: readxls2list_findheader
    # the function name: def excelDict2list_findheader(
    # def test_excelDict2list_findheader_p01_pass(self):
    ########################################

    ### READXLS_FINDHEADER

    ########################################
    # the function name: def readxls_findheader(xlsfile, req_cols, xlatdict=None, optiondict=None, col_aref=None, data_only=False, debug=False):
    # XLS/XLSX - simple open and return excel_dict
    def test_readxls_findheader_p01_xls_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexls, req_cols, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 0)
        self.assertEqual(excel_dict["sheetmaxrow"], 12)

    def test_readxls_findheader_p02_xlsx_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 0)
        self.assertEqual(excel_dict["sheetmaxrow"], 12)

    # XLS type of field check
    def test_readxls_findheader_f01_xls_fld_type_col_aref(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexls, req_cols, col_aref={}, debug=False
            )

    def test_readxls_findheader_f02_xls_fld_type_req_cols(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(filenamexls, {}, debug=False)

    def test_readxls_findheader_f03_xls_fld_type_optiondict(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexls, req_cols, optiondict="", debug=False
            )

    def test_readxls_findheader_f03_xls_fld_type_xlatdict(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexls, req_cols, xlatdict="", debug=False
            )

    # XLSX type of field check
    def test_readxls_findheader_f01_xlsx_fld_type_col_aref(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols, col_aref={}, debug=False
            )

    def test_readxls_findheader_f02_xlsx_fld_type_req_cols(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(filenamexlsx, {}, debug=False)

    def test_readxls_findheader_f03_xlsx_fld_type_optiondict(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols, optiondict="", debug=False
            )

    def test_readxls_findheader_f03_xlsx_fld_type_xlatdict(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols, xlatdict="", debug=False
            )

    # xlatdict
    def test_readxls_findheader_p01_xls_xlat_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexls, req_cols_xlat, xlatdict=xlatdict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 0)

    def test_readxls_findheader_p02_xlsx_xlat_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols_xlat, xlatdict=xlatdict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 0)

    def test_readxls_findheader_f01_xls_xlat_pass(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexls, req_cols2, xlatdict=xlatdict, debug=False
            )

    def test_readxls_findheader_f02_xlsx_xlat_pass(self):
        with self.assertRaises(Exception) as context:
            excel_dict = kvxls.readxls_findheader(
                filenamexlsx, req_cols2, xlatdict=xlatdict, debug=False
            )

    # col_header
    def test_readxls_findheader_p01_xls_col_header(self):
        optiondict = {"col_header": True}
        excel_dict = kvxls.readxls_findheader(
            filenamexls, [], optiondict=optiondict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 0)

    def test_readxls_findheader_p02_xlsx_col_header(self):
        optiondict = {"col_header": True}
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, [], optiondict=optiondict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 0)

    # no_header
    def test_readxls_findheader_p01_xls_no_header(self):
        optiondict = {"no_header": True}
        col_aref = list(records[0].keys())
        excel_dict = kvxls.readxls_findheader(
            filenamexls,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], None)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 0)

    def test_readxls_findheader_p02_xlsx_no_header(self):
        optiondict = {"no_header": True}
        col_aref = list(records[0].keys())
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], None)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 0)

    # col_aref only
    def test_readxls_findheader_p01_xls_col_aref(self):
        optiondict = {}
        col_aref = ["ken1", "ken2", "ken3", "ken4"]
        excel_dict = kvxls.readxls_findheader(
            filenamexls,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 0)
        self.assertEqual(
            excel_dict["header"],
            [
                "ken1",
                "ken2",
                "ken3",
                "ken4",
                "blank001",
                "blank002",
                "blank003",
            ],
        )

    def test_readxls_findheader_p01_xlsx_col_aref(self):
        optiondict = {}
        col_aref = ["ken1", "ken2", "ken3", "ken4"]
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 0)
        self.assertEqual(
            excel_dict["header"],
            [
                "ken1",
                "ken2",
                "ken3",
                "ken4",
                "blank001",
                "blank002",
                "blank003",
            ],
        )

    # save_row
    def test_readxls_findheader_p01_xls_save_row(self):
        optiondict = {"save_row": True}
        excel_dict = kvxls.readxls_findheader(
            filenamexls, [], optiondict=optiondict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 0)
        # need to define what we do with row header on - nothing today
        # self.assertEqual(excel_dict['header'][-1], 'XLSRow')

    def test_readxls_findheader_p01_xlsx_save_row(self):
        optiondict = {"save_row": True}
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, [], optiondict=optiondict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 0)
        # need to define what we do with row header on - nothing today
        # self.assertEqual(excel_dict['header'][-1], 'XLSRow')

    # simple put set the start row
    def test_readxls_findheader_p01_xls_start_row_col_aref_blank_req_cols(self):
        optiondict = {"start_row": 3}
        col_aref = ["ken1", "ken2", "ken3", "ken4"]
        excel_dict = kvxls.readxls_findheader(
            filenamexls,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], None)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 2)
        self.assertEqual(
            excel_dict["header"],
            [
                "ken1",
                "ken2",
                "ken3",
                "ken4",
                "blank001",
                "blank002",
                "blank003",
            ],
        )

    def test_readxls_findheader_p02_xlsx_start_row_col_aref_pass(self):
        optiondict = {"start_row": 3}
        col_aref = ["ken1", "ken2", "ken3", "ken4"]
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], None)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 2)
        self.assertEqual(
            excel_dict["header"],
            [
                "ken1",
                "ken2",
                "ken3",
                "ken4",
                "blank001",
                "blank002",
                "blank003",
            ],
        )

    # maxrowls
    def test_readxls_findheader_p01_xls_max_rows(self):
        optiondict = {"max_rows": 6}
        excel_dict = kvxls.readxls_findheader(
            filenamexls, req_cols, optiondict=optiondict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexls)
        self.assertEqual(excel_dict["start_row"], 0)
        self.assertEqual(excel_dict["sheetmaxrow"], 6)

    def test_readxls_findheader_p02_xlsx_max_rows(self):
        optiondict = {"max_rows": 6}
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, optiondict=optiondict, debug=False
        )
        self.assertEqual(type(excel_dict), dict)
        self.assertEqual(excel_dict["keep_vba"], True)
        self.assertEqual(excel_dict["row_header"], 0)
        self.assertEqual(excel_dict["xlsfile"], filenamexlsx)
        self.assertEqual(excel_dict["start_row"], 0)
        self.assertEqual(excel_dict["sheetmaxrow"], 6)

    ### CHGSHEET_FINDHEADER

    ########################################
    # the function name: def chgsheet_findheader(excel_dict, req_cols, xlatdict=None, optiondict=None,
    def test_chgsheet_findheader_p01_xls_pass(self):
        # read in teh first time
        excel_dict41 = kvxls.readxls_findheader(
            filenamexls4, req_cols4_1, optiondict=optiondict41, debug=False
        )
        # change sheet to the sheet of interest
        excel_dict42 = kvxls.chgsheet_findheader(
            excel_dict41, req_cols4_2, optiondict=optiondict42, debug=False
        )
        self.assertEqual(type(excel_dict42), dict)
        self.assertEqual(excel_dict42["keep_vba"], True)
        self.assertEqual(excel_dict42["row_header"], 0)
        self.assertEqual(excel_dict42["xlsfile"], filenamexls4)
        self.assertEqual(excel_dict42["sheetmaxrow"], 3)
        self.assertEqual(excel_dict42["sheet_name"], optiondict42["sheet_name"])

    def test_chgsheet_findheader_p02_xlsx_pass(self):
        # read in teh first time
        excel_dict41 = kvxls.readxls_findheader(
            filenamexlsx4, req_cols4_1, optiondict=optiondict41, debug=False
        )
        # change sheet to the sheet of interest
        excel_dict42 = kvxls.chgsheet_findheader(
            excel_dict41, req_cols4_2, optiondict=optiondict42, debug=False
        )
        self.assertEqual(type(excel_dict42), dict)
        self.assertEqual(excel_dict42["keep_vba"], True)
        self.assertEqual(excel_dict42["row_header"], 0)
        self.assertEqual(excel_dict42["xlsfile"], filenamexlsx4)
        self.assertEqual(excel_dict42["start_row"], 0)
        self.assertEqual(excel_dict42["sheetmaxrow"], 3)
        self.assertEqual(excel_dict42["sheet_name"], optiondict42["sheet_name"])

    ### READXLS2LIST_FINDHEADER

    # XLS file processing - simple req_cols
    def test_readxls2list_findheader_p01_xls_simple_reqcols(self):
        result = kvxls.readxls2list_findheader(
            filenamexls, req_cols, debug=False
        )
        self.assertEqual(result[0], records[0])
        self.assertEqual(len(result), len(records))

    def test_readxls2list_findheader_p02_xls_simple_reqcols_aref_result(self):
        result = kvxls.readxls2list_findheader(
            filenamexls, req_cols, optiondict={"aref_result": True}, debug=False
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p03_xls_simple_reqcols_col_header(self):
        result = kvxls.readxls2list_findheader(
            filenamexls, req_cols, optiondict={"col_header": True}, debug=False
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p04_xls_simple_reqcols_no_header_start_row_return_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls,
            req_cols,
            optiondict={"no_header": True, "start_row": 2},
            debug=False,
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p05_xls_simple_reqcols_no_header_start_row_col_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls,
            req_cols,
            optiondict={"no_header": True, "start_row": 2},
            col_aref=list(records[0].keys()),
            debug=False,
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p06_xls_simple_reqcols_no_header_start_row_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexls,
            req_cols,
            optiondict={"no_header": True, "start_row": 2},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p07_xls_simple_reqcols_col_aref_blank_column(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexls,
            req_cols,
            optiondict={"save_row": True},
            col_aref=col_aref,
            debug=False,
        )
        self.assertTrue("blank001" in result[0].keys())

    def test_readxls2list_findheader_p08_xls_simple_reqcols_col_header_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexls,
            req_cols,
            optiondict={"col_header": True},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p09_xls_simple_reqcols_convert_dateflds(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls2,
            req_cols2,
            optiondict={"dateflds": ["DateField"]},
            debug=False,
        )
        self.assertEqual(result[0], records2[0])

    def test_readxls2list_findheader_p10_xls_simple_reqcols_save_row(self):
        result = kvxls.readxls2list_findheader(
            filenamexls, req_cols, optiondict={"save_row": True}, debug=False
        )
        self.assertEqual(result[0]["XLSRow"], 2)

    def test_readxls2list_findheader_p11_xls_simple_reqcols_aref_result_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexls3, aref, optiondict={"aref_result": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexls3,
            req_cols,
            optiondict={"aref_result": True},
            debug=False,
        )
        self.assertEqual(result[0], list(records[0].values()))
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_p12_xls_simple_reqcol_no_header_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexls3, aref, optiondict={"no_header": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexls3, req_cols, optiondict={"no_header": True}, debug=False
        )
        self.assertEqual(result[0], ["", "", "", "", "", "", ""])
        self.assertEqual(result[5], list(records[0].keys()))
        self.assertEqual(result[6], list(records[0].values()))
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_p13_xls_simple_reqcols_break_blank_row(
        self,
    ):
        # get the first 6 records
        aref = records[:6]
        # create a non record
        rec = {x: None for x in records[0].keys()}
        # make 2 blank records
        aref.append(rec)
        aref.append(rec)
        # add two no blank records
        aref.extend(records[6:8])
        # save this out
        kvxls.writelist2xls(filenamexls3, aref, debug=False)
        # now read in the file with break on blank lines
        result = kvxls.readxls2list_findheader(
            filenamexls3,
            req_cols,
            optiondict={"break_blank_row": True},
            debug=False,
        )
        self.assertEqual(len(result), 6)
        # now read in the file with OUT break on blank lines
        result = kvxls.readxls2list_findheader(
            filenamexls3, req_cols, debug=False
        )
        self.assertEqual(len(result), 10)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_p14_xls_simple_reqcols_skip_blank_row(
        self,
    ):
        # get the first 6 records
        aref = records[:6]
        # create a non record
        rec = {x: None for x in records[0].keys()}
        # make 2 blank records
        aref.append(rec)
        aref.append(rec)
        # add two no blank records
        aref.extend(records[6:8])
        # save this out
        kvxls.writelist2xls(filenamexls3, aref, debug=False)
        # now read in the file with break on blank lines
        result = kvxls.readxls2list_findheader(
            filenamexls3,
            req_cols,
            optiondict={"skip_blank_row": True},
            debug=False,
        )
        self.assertEqual(len(result), 8)
        # now read in the file with OUT break on blank lines
        result = kvxls.readxls2list_findheader(
            filenamexls3, req_cols, debug=False
        )
        self.assertEqual(len(result), 10)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_p15_xls_simple_col_header_start_row_no_req_cols(
        self,
    ):
        kvxls.writelist2xls(
            filenamexls3, records, optiondict={"start_row": 3}, debug=False
        )
        result = kvxls.readxls2list_findheader(
            filenamexls3,
            [],
            optiondict={"col_header": True, "start_row": 3},
            debug=False,
        )
        self.assertEqual(result[0], records[0])
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_f01_xls_simple_col_header_start_row_no_req_cols_invalid_start_row(
        self,
    ):
        kvxls.writelist2xls(
            filenamexls3, records, optiondict={"start_row": 3}, debug=False
        )
        with self.assertRaises(Exception) as context:
            result = kvxls.readxls2list_findheader(
                filenamexls3,
                [],
                optiondict={"col_header": True, "start_row": 2},
                debug=False,
            )
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    # XLS file processing - simple blank req_cols
    def test_readxls2list_findheader_p01_xls_simple_blankReqCols(self):
        result = kvxls.readxls2list_findheader(filenamexls, [], debug=False)
        self.assertEqual(result[0], records[0])
        self.assertEqual(len(result), len(records))

    def test_readxls2list_findheader_p02_xls_simple_blankReqCols_aref_result(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls, [], optiondict={"aref_result": True}, debug=False
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p03_xls_simple_blankReqCols_col_header(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls, [], optiondict={"col_header": True}, debug=False
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p04_xls_simple_blankReqCols_no_header_start_row_return_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls,
            [],
            optiondict={"no_header": True, "start_row": 2},
            debug=False,
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p05_xls_simple_blankReqCols_no_header_start_row_col_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls,
            [],
            optiondict={"no_header": True, "start_row": 2},
            col_aref=list(records[0].keys()),
            debug=False,
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p06_xls_simple_blankReqCols_no_header_start_row_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexls,
            [],
            optiondict={"no_header": True, "start_row": 2},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p07_xls_simple_blankReqCols_col_aref_blank_column(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexls,
            [],
            optiondict={"save_row": True},
            col_aref=col_aref,
            debug=False,
        )
        self.assertTrue("blank001" in result[0].keys())

    def test_readxls2list_findheader_p08_xls_simple_blankReqCols_col_header_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexls,
            [],
            optiondict={"col_header": True},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p09_xls_simple_blankReqCols_convert_dateflds(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexls2,
            [],
            optiondict={"dateflds": ["DateField"]},
            debug=False,
        )
        self.assertEqual(result[0], records2[0])

    def test_readxls2list_findheader_p10_xls_simple_blankReqCols_save_row(self):
        result = kvxls.readxls2list_findheader(
            filenamexls, [], optiondict={"save_row": True}, debug=False
        )
        self.assertEqual(result[0]["XLSRow"], 2)

    def test_readxls2list_findheader_p11_xls_simple_blankReqCols_aref_result_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexls3, aref, optiondict={"aref_result": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexls3, [], optiondict={"aref_result": True}, debug=False
        )
        self.assertEqual(result[0], ["", "", "", "", "", "", ""])
        self.assertEqual(result[5], list(records[0].values()))
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_p12_xls_simple_blankReqCols_no_header_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexls3, aref, optiondict={"no_header": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexls3, [], optiondict={"no_header": True}, debug=False
        )
        self.assertEqual(result[0], ["", "", "", "", "", "", ""])
        self.assertEqual(result[5], list(records[0].keys()))
        self.assertEqual(result[6], list(records[0].values()))
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_f01_xls_simple_maxrows_exceeded_in_header_search(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexls3, aref, optiondict={"no_header": True}, debug=False
        )
        with self.assertRaises(Exception) as context:
            pprint.pprint(
                kvxls.readxls2list_findheader(
                    filenamexls3,
                    req_cols,
                    optiondict={"maxrows": 2},
                    debug=False,
                )
            )
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_readxls2list_findheader_f02_xls_simple_unique_columns_test(self):
        header = list(records[0].keys())
        dupkey = header[0]
        header.append(dupkey)
        aref = []
        for rec in records:
            aref.append(list(rec.values()) + [rec[dupkey]])
        kvxls.writelist2xls(
            filenamexls3, aref, optiondict={"aref_result": True}, debug=False
        )
        with self.assertRaises(Exception) as context:
            kvxls.readxls2list_findheader(
                filenamexls3,
                req_cols,
                optiondict={"unique_column": True},
                debug=False,
            )
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    # XLSX file processing
    def test_readxls2list_findheader_p01_xlsx_simple_reqcols(self):
        result = kvxls.readxls2list_findheader(
            filenamexlsx, req_cols, debug=False
        )
        self.assertEqual(result[0], records[0])
        self.assertEqual(len(result), len(records))

    def test_readxls2list_findheader_p02_xlsx_simple_reqcols_aref_result(self):
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            req_cols,
            optiondict={"aref_result": True},
            debug=False,
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p03_xlsx_simple_reqcols_col_header(self):
        result = kvxls.readxls2list_findheader(
            filenamexlsx, req_cols, optiondict={"col_header": True}, debug=False
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p04_xlsx_simple_reqcols_no_header_start_row_return_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            req_cols,
            optiondict={"no_header": True, "start_row": 2},
            debug=False,
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p05_xlsx_simple_reqcols_no_header_start_row_col_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            req_cols,
            optiondict={"no_header": True, "start_row": 2},
            col_aref=list(records[0].keys()),
            debug=False,
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p06_xlsx_simple_reqcols_no_header_start_row_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            req_cols,
            optiondict={"no_header": True, "start_row": 2},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p07_xlsx_simple_reqcols_col_aref_blank_column(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            req_cols,
            optiondict={"save_row": True},
            col_aref=col_aref,
            debug=False,
        )
        self.assertTrue("blank001" in result[0].keys())

    def test_readxls2list_findheader_p08_xlsx_simple_reqcols_col_header_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            req_cols,
            optiondict={"col_header": True},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p09_xlsx_simple_reqcols_convert_dateflds(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx2,
            req_cols2,
            optiondict={"dateflds": ["DateField"]},
            debug=False,
        )
        self.assertEqual(result[0], records2[0])

    def test_readxls2list_findheader_p10_xlsx_simple_reqcols_save_row(self):
        result = kvxls.readxls2list_findheader(
            filenamexlsx, req_cols, optiondict={"save_row": True}, debug=False
        )
        self.assertEqual(result[0]["XLSRow"], 2)

    def test_readxls2list_findheader_p11_xlsx_simple_reqcols_aref_result_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexlsx3, aref, optiondict={"aref_result": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexlsx3,
            req_cols,
            optiondict={"aref_result": True},
            debug=False,
        )
        self.assertEqual(result[0], list(records[0].values()))
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_readxls2list_findheader_p12_xlsx_simple_reqcol_no_header_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexlsx3, aref, optiondict={"no_header": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexlsx3, req_cols, optiondict={"no_header": True}, debug=False
        )
        self.assertEqual(result[0], [None, None, None, None, None, None, None])
        self.assertEqual(result[5], list(records[0].keys()))
        self.assertEqual(result[6], list(records[0].values()))
        # self.assertEqual(result[0], records[0])
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    # XLSX file processing - simple blank req_cols
    def test_readxls2list_findheader_p01_xlsx_simple_blankReqCols(self):
        result = kvxls.readxls2list_findheader(filenamexlsx, [], debug=False)
        self.assertEqual(result[0], records[0])
        self.assertEqual(len(result), len(records))

    def test_readxls2list_findheader_p02_xlsx_simple_blankReqCols_aref_result(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx, [], optiondict={"aref_result": True}, debug=False
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p03_xlsx_simple_blankReqCols_col_header(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx, [], optiondict={"col_header": True}, debug=False
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p04_xlsx_simple_blankReqCols_no_header_start_row_return_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            [],
            optiondict={"no_header": True, "start_row": 2},
            debug=False,
        )
        self.assertEqual(result[0], list(records[0].values()))

    def test_readxls2list_findheader_p05_xlsx_simple_blankReqCols_no_header_start_row_col_aref(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            [],
            optiondict={"no_header": True, "start_row": 2},
            col_aref=list(records[0].keys()),
            debug=False,
        )
        self.assertEqual(result[0], records[0])

    def test_readxls2list_findheader_p06_xlsx_simple_blankReqCols_no_header_start_row_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            [],
            optiondict={"no_header": True, "start_row": 2},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p07_xlsx_simple_blankReqCols_col_aref_blank_column(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            [],
            optiondict={"save_row": True},
            col_aref=col_aref,
            debug=False,
        )
        self.assertTrue("blank001" in result[0].keys())

    def test_readxls2list_findheader_p08_xlsx_simple_blankReqCols_col_header_col_aref_missing_cols(
        self,
    ):
        col_aref = list(records[0].keys())[:-2]
        result = kvxls.readxls2list_findheader(
            filenamexlsx,
            [],
            optiondict={"col_header": True},
            col_aref=col_aref,
            debug=False,
        )
        # manipulate the data to make them match - first copy the data so we don't change the original
        temprec = dict(records[0])
        temprec["blank001"] = temprec["Type"]
        del temprec["Type"]
        temprec["blank002"] = temprec["LastSeen"]
        del temprec["LastSeen"]
        self.assertEqual(result[0], temprec)

    def test_readxls2list_findheader_p09_xlsx_simple_blankReqCols_convert_dateflds(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx2,
            [],
            optiondict={"dateflds": ["DateField"]},
            debug=False,
        )
        self.assertEqual(result[0], records2[0])

    def test_readxls2list_findheader_p10_xlsx_simple_blankReqCols_save_row(
        self,
    ):
        result = kvxls.readxls2list_findheader(
            filenamexlsx, [], optiondict={"save_row": True}, debug=False
        )
        self.assertEqual(result[0]["XLSRow"], 2)

    def test_readxls2list_findheader_p11_xlsx_simple_blankReqCols_aref_result_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexlsx3, aref, optiondict={"aref_result": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexlsx3, [], optiondict={"aref_result": True}, debug=False
        )
        self.assertEqual(result[0], [None, None, None, None, None, None, None])
        self.assertEqual(result[5], list(records[0].values()))
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_readxls2list_findheader_p12_xlsx_simple_blankReqCols_no_header_starting_blank_lines(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexlsx3, aref, optiondict={"no_header": True}, debug=False
        )
        # now read in the file
        result = kvxls.readxls2list_findheader(
            filenamexlsx3, [], optiondict={"no_header": True}, debug=False
        )
        self.assertEqual(result[0], [None, None, None, None, None, None, None])
        self.assertEqual(result[5], list(records[0].keys()))
        self.assertEqual(result[6], list(records[0].values()))
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_readxls2list_findheader_f01_xlsx_maxrows_exceeded_in_header_search(
        self,
    ):
        # create a list of values that are used to create the xls - we have 6 blank lines at the top
        aref = [[""], [""], [""], [""], [""]]
        aref.append(list(records[0].keys()))
        for rec in records:
            aref.append(list(rec.values()))
        kvxls.writelist2xls(
            filenamexlsx3, aref, optiondict={"no_header": True}, debug=False
        )
        with self.assertRaises(Exception) as context:
            kvxls.readxls2list_findheader(
                filenamexlsx3, req_cols, optiondict={"maxrows": 2}, debug=False
            )
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_readxls2list_findheader_f02_xlsx_unique_columns_test(self):
        header = list(records[0].keys())
        dupkey = header[0]
        header.append(dupkey)
        aref = []
        for rec in records:
            aref.append(list(rec.values()) + [rec[dupkey]])
        kvxls.writelist2xls(
            filenamexlsx3, aref, optiondict={"aref_result": True}, debug=False
        )
        with self.assertRaises(Exception) as context:
            kvxls.readxls2list_findheader(
                filenamexlsx3,
                req_cols,
                optiondict={"unique_column": True},
                debug=False,
            )
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    ########################################
    # the function name: def readxls2dict_findheader(xlsfile, dictkeys, req_cols=None, xlatdict=None, optiondict=None,
    def test_readxls2dict_findheader_p01_xls_pass(self):
        results = kvxls.readxls2dict_findheader(
            filenamexls, req_cols, req_cols, debug=False
        )
        self.assertEqual(list(results.keys()), business_keys)

    def test_readxls2dict_findheader_p02_xlsx_pass(self):
        results = kvxls.readxls2dict_findheader(
            filenamexlsx, req_cols, req_cols, debug=False
        )
        self.assertEqual(list(results.keys()), business_keys)

    def test_readxls2dict_findheader_p03_xlsx_req_cols_string(self):
        results = kvxls.readxls2dict_findheader(
            filenamexlsx, "Company", req_cols, debug=False
        )
        self.assertEqual(
            list(results.keys()),
            ["NHLiq", "BevMo", "WineClub", "TotalCA", "Vons"],
        )

    def test_readxls2dict_findheader_f01_xlsx_dictkeys_none(self):
        with self.assertRaises(Exception) as context:
            kvxls.readxls2dict_findheader(
                filenamexlsx, None, req_cols, debug=False
            )

    def test_readxls2dict_findheader_f02_xlsx_dictkeys_int(self):
        with self.assertRaises(Exception) as context:
            kvxls.readxls2dict_findheader(
                filenamexlsx, 1, req_cols, debug=False
            )

    def test_readxls2dict_findheader_f03_xlsx_req_cols_int(self):
        with self.assertRaises(Exception) as context:
            kvxls.readxls2dict_findheader(
                filenamexlsx, req_cols, 1, debug=False
            )

    ########################################
    # the function name: def writedict2xls(xlsfile, data, col_aref=None, optiondict={}, debug=False):
    def test_writedict2xls_p01_xlsx_simple_pass(self):
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2dict_findheader(
            filenamexlsx, req_cols, req_cols, debug=False
        )
        filename = kvxls.writedict2xls(filenamexlsx4, results, debug=False)
        results2 = kvxls.readxls2dict_findheader(
            filenamexlsx4, req_cols, req_cols, debug=False
        )
        self.assertEqual(results, results2)
        self.assertEqual(filename, filenamexlsx4)
        kvutil.remove_filename(filenamexlsx4, kvutil.functionName())

    ########################################
    # the function name: def writelist2xls(xlsfile, data, col_aref=None, optiondict=None, debug=False):
    def test_writelist2xls_p01_xlsx_simple_pass(self):
        # list of dicts
        filename = kvxls.writelist2xls(filenamexlsx3, records, debug=False)
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2list(filenamexlsx3)
        self.assertEqual(records, results)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p02_xlsx_nodata_allow_empty(self):
        # no data
        filename = kvxls.writelist2xls(filenamexlsx3, None, debug=False)
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2list(
            filenamexlsx3, optiondict={"allow_empty": True}, debug=False
        )
        self.assertEqual(results, [])
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p03_xlsx_set_sheet(self):
        # list of dicts and set the sheet name created
        optiondict = {"sheet_name": "set_sheet_name"}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2list(filenamexlsx3, optiondict=optiondict)
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p04_xlsx_set_and_limit_column_order(self):
        # list of dicts - set column order and limit columns
        col_aref = ["Wine", "Company", "Vintage"]
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, col_aref=col_aref, debug=False
        )
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2list(filenamexlsx3)
        match_records = [{x: v[x] for x in col_aref} for v in records]
        self.assertEqual(results, match_records)
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p05_xlsx_list(self):
        # array of data and we pass in the header
        rec_array = [[v[k] for k in v.keys()] for v in records]
        filename = kvxls.writelist2xls(
            filenamexlsx3,
            rec_array,
            col_aref=list(records[0].keys()),
            debug=False,
        )
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2list(filenamexlsx3)
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p06_xlsx_set_sheet_2sheets(self):
        # list of dicts and set the sheet name created
        optiondict = {"sheet_name": "set_sheet_name1"}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2list(filenamexlsx3, optiondict=optiondict)
        self.assertEqual(records, results)
        # now build 2nd sheet
        optiondict2 = {"sheet_name": "set_sheet_name2", "replace_sheet": True}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict2, debug=False
        )
        results = kvxls.readxls2list(filenamexlsx3, optiondict=optiondict)
        self.assertEqual(records, results)
        results = kvxls.readxls2list(filenamexlsx3, optiondict=optiondict2)
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p07_xlsx_aref_result(self):
        # array of data and we pass in the header
        rec_array = [[v[k] for k in v.keys()] for v in records]
        optiondict = {"aref_result": True}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexlsx3)
        results = kvxls.readxls2list(filenamexlsx3, optiondict=optiondict)
        self.assertEqual(rec_array, results)
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p08_xlsx_no_header(self):
        # array of data and we pass in the header
        col_aref = list(records[0].keys())
        optiondict = {"no_header": True}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexlsx3)
        #        results = kvxls.readxls2list( filenamexlsx3, optiondict=optiondict, col_aref=col_aref )
        results = kvxls.readxls2list_findheader(
            filenamexlsx3,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_p09_xlsx_start_row(self):
        # list of dicts
        optiondict = {"start_row": 3}
        filename = kvxls.writelist2xls(
            filenamexlsx3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexlsx3)
        optiondict = {"no_header": True}
        col_aref = None
        results = kvxls.readxls2list_findheader(
            filenamexlsx3,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(results[0], [None, None, None, None, None, None, None])
        self.assertEqual(results[2], list(records[0].keys()))
        self.assertEqual(results[3], list(records[0].values()))
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    def test_writelist2xls_f01_xlsx_nodata_not_allow_empty(self):
        # no data
        filename = kvxls.writelist2xls(filenamexlsx3, None, debug=False)
        # self.assertEqual(filename, filenamexlsx3)
        with self.assertRaises(Exception) as context:
            results = kvxls.readxls2list(filenamexlsx3)
        self.assertEqual(filename, filenamexlsx3)
        kvutil.remove_filename(filenamexlsx3, kvutil.functionName())

    ########################################
    # the function name: def writelist2xls(xlsfile, data, col_aref=None, optiondict=None, debug=False):
    def test_writelist2xls_p01_xls_simple_pass(self):
        # list of dicts
        filename = kvxls.writelist2xls(filenamexls3, records, debug=False)
        # self.assertEqual(filename, filenamexls3)
        results = kvxls.readxls2list(filenamexls3)
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p02_xls_nodata_allow_empty(self):
        # no data
        filename = kvxls.writelist2xls(filenamexls3, None, debug=False)
        # self.assertEqual(filename, filenamexls3)
        results = kvxls.readxls2list(
            filenamexls3, optiondict={"allow_empty": True}, debug=False
        )
        self.assertEqual(results, [])
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p03_xls_set_sheet(self):
        # list of dicts and set the sheet name created
        optiondict = {"sheet_name": "set_sheet_name"}
        filename = kvxls.writelist2xls(
            filenamexls3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexls3)
        results = kvxls.readxls2list(filenamexls3, optiondict=optiondict)
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p04_xls_set_and_limit_column_order(self):
        # list of dicts - set column order and limit columns
        col_aref = ["Wine", "Company", "Vintage"]
        filename = kvxls.writelist2xls(
            filenamexls3, records, col_aref=col_aref, debug=False
        )
        # self.assertEqual(filename, filenamexls3)
        results = kvxls.readxls2list(filenamexls3)
        match_records = [{x: v[x] for x in col_aref} for v in records]
        self.assertEqual(results, match_records)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p05_xls_list(self):
        # array of data and we pass in the header
        rec_array = [[v[k] for k in v.keys()] for v in records]
        filename = kvxls.writelist2xls(
            filenamexls3,
            rec_array,
            col_aref=list(records[0].keys()),
            debug=False,
        )
        # self.assertEqual(filename, filenamexls3)
        results = kvxls.readxls2list(filenamexls3)
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p06_xls_set_sheet_2sheets(self):
        # list of dicts and set the sheet name created
        optiondict = {"sheet_name": "set_sheet_name1"}
        filename = kvxls.writelist2xls(
            filenamexls3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexls3)
        results = kvxls.readxls2list(filenamexls3, optiondict=optiondict)
        self.assertEqual(records, results)
        # now build 2nd sheet
        optiondict2 = {"sheet_name": "set_sheet_name2", "replace_sheet": True}
        filename = kvxls.writelist2xls(
            filenamexls3, records, optiondict=optiondict2, debug=False
        )
        results = kvxls.readxls2list(filenamexls3, optiondict=optiondict)
        self.assertEqual(records, results)
        results = kvxls.readxls2list(filenamexls3, optiondict=optiondict2)
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p07_xls_aref_result(self):
        # array of data and we pass in the header
        rec_array = [[v[k] for k in v.keys()] for v in records]
        optiondict = {"aref_result": True}
        filename = kvxls.writelist2xls(
            filenamexls3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexls3)
        results = kvxls.readxls2list(filenamexls3, optiondict=optiondict)
        self.assertEqual(rec_array, results)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p08_xls_no_header(self):
        # array of data and we pass in the header
        col_aref = list(records[0].keys())
        optiondict = {"no_header": True}
        filename = kvxls.writelist2xls(
            filenamexls3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexls3)
        #        results = kvxls.readxls2list( filenamexls3, optiondict=optiondict, col_aref=col_aref )
        results = kvxls.readxls2list_findheader(
            filenamexls3,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(records, results)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_p09_xls_start_row(self):
        # list of dicts
        optiondict = {"start_row": 3}
        filename = kvxls.writelist2xls(
            filenamexls3, records, optiondict=optiondict, debug=False
        )
        # self.assertEqual(filename, filenamexls3)
        optiondict = {"no_header": True}
        col_aref = None
        results = kvxls.readxls2list_findheader(
            filenamexls3,
            [],
            optiondict=optiondict,
            col_aref=col_aref,
            debug=False,
        )
        self.assertEqual(results[0], ["", "", "", "", "", "", ""])
        self.assertEqual(results[2], list(records[0].keys()))
        self.assertEqual(results[3], list(records[0].values()))
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    def test_writelist2xls_f01_xls_nodata_not_allow_empty(self):
        # no data
        filename = kvxls.writelist2xls(filenamexls3, None, debug=False)
        # self.assertEqual(filename, filenamexls3)
        with self.assertRaises(Exception) as context:
            results = kvxls.readxls2list(filenamexls3, debug=False)
        self.assertEqual(filename, filenamexls3)
        kvutil.remove_filename(filenamexls3, kvutil.functionName())

    # the function name: def writexls(excel_dict, xlsfile, xlsm=False, debug=False):
    def test_writexls_p01_pass(self):
        excel_dict = kvxls.readxls_findheader(
            filenamexlsx, req_cols, debug=False
        )
        kvxls.writexls(excel_dict)


if __name__ == "__main__":
    unittest.main()
